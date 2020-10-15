# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl

original_filename = r"C:\Users\Root\Desktop\test\test.xlsx"
new_filename = r"C:\Users\Root\Desktop\test\test_good.xlsx"
def prepare_df(filename):
    df = pd.read_excel(filename, header=3, engine='openpyxl')
#     df = df.dropna(how='all')
    df['util'] = df['1'].fillna(method='ffill')
#     df = df.fillna('')
    return df
 
df = prepare_df(original_filename)
 
def write_dataframe(df, filename, limit=100, sheet_name_base='Лист '):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        items = df['util'].unique().tolist()
        lst_num = 0
        buffer = []
        while items:
            buffer.append(items.pop(0))
            current_part = df[df['util'].isin(buffer)]
            if current_part.shape[0] <= limit:
                continue
            else:
                current_part = df[df['util'].isin(buffer[:-1])].drop('util', axis='columns')
                current_part.to_excel(writer, startrow=3, index=False, sheet_name=f"{sheet_name_base}{lst_num}")
                lst_num += 1
                buffer = buffer[-1:]
        df[df['util'].isin(buffer)].drop('util', axis='columns').to_excel(writer, startrow=3, index=False, sheet_name=f"{sheet_name_base}{lst_num}")
 
write_dataframe(df.iloc[:500, :], new_filename, limit=100)
 
def copy_header(header, sheet):
    for row in header:
        for cell in row:
            new_cell = sheet.cell(row=cell.row, column=cell.column, value=cell.value)
 
def fill_headers(from_filename, to_filename):
    header = openpyxl.load_workbook(from_filename).active[1:3]
    wb = openpyxl.load_workbook(to_filename)
    for sheet in wb:
        copy_header(header, sheet)
    wb.save(to_filename)
 
fill_headers(original_filename, new_filename)